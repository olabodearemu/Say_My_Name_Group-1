import pyttsx3 as texttospeech;       #For converting text to speech
import openpyxl as excel;             #For reading the excel file
import eng_to_ipa as phonetic;        #For getting the phonetic spelling of the text
from g2p_en import G2p;
import nltk

#Reading the excel file 
students = r'C:\Users\nehap\OneDrive\Desktop\AIGS\Machine Learning\ML_Project\SayMyName_Backend\Student_details.xlsx'
wb_obj = excel.load_workbook(students)
sheet = wb_obj['Sheet1']
row = sheet.max_row         #Stores the max value of the row
column = sheet.max_column   #Stores the max value of the column

#Prompting the user to enter their Full Name
user_name = input("Please enter your first name:")

#Initializing text to speech 
tts = texttospeech.init()

#g2p initialization
g2p = G2p()

# Download the cmudict corpus that contains the pronunciation of words
#nltk.download('cmudict')
 
# Load the cmudict dictionary
cmudict = nltk.corpus.cmudict.dict()

for i in range(1, row + 1):  
    cell_name = sheet.cell(row = i+1, column = 1)

    #This block of code executes if the input name exists in the Excel
    if cell_name.value.lower() == user_name.lower():
        print(f"The Name \"{cell_name.value.lower()}\" exists in the Database")

        # using g2p for generating phonetics
        phonetic_name = g2p(cell_name.value.lower())

        # appending the string from the list generated above of phonetics
        word = ''
        for x in phonetic_name:
            word = word + x
     
        # taking out the number from the string and then joining the string by -
        true_phonetices = ''.join("-" if c.isdigit() else c for c in word)
        if true_phonetices[len(true_phonetices)-1] == "-":
            true_phonetices = true_phonetices[:-1] + ""
        print(true_phonetices.lower())
    
        tts.say(user_name.lower())      #Say function is used to speak the name which the User has entered
        tts.runAndWait()

        break

    #This block of code executes if the input name does not exist in the Excel
    else:
        i = i + 1       #Iterates over each row to check if the name is present
        
        #This block of code executes once we reached the last row and we didn't find the name
        if i == row:    
            print("No such Name in the Database.")
            student_id = int(input("Please enter your student id: "))

            #This loop executes and checks for the 2nd column which is the student ID
            for j in range(1, row+1):
                cell_id = sheet.cell(row = j+1, column = 2)

                #This block of code executes if the entered student ID matches with any other student's id
                if cell_id.value == student_id:
                    print(f"The ID {cell_id.value} exists for other Student.")
                    break
                #This block of code will store the Name and Student ID if it is not already present in the Excel and will speak the Name
                else:
                    j = j + 1
                    if j == row:
                        sheet.append([user_name, student_id])
                        wb_obj.save(students)
                        i = i + 1
                        j = j + 1
                        row = sheet.max_row 
                        cell_name = sheet.cell(row = i, column = 1)
                       
                        # using g2p for generating phonetics
                        phonetic_name = g2p(cell_name.value.lower())

                        # appending the string from the list generated above of phonetics
                        word = ''
                        for x in phonetic_name:
                            word = word + x

                        # taking out number from the string
                        true_phonetices = ''.join("-" if c.isdigit() else c for c in word)
                        if true_phonetices[len(true_phonetices)-1] == "-":
                            true_phonetices = true_phonetices[:-1] + ""
                        print(true_phonetices.lower())
                        cell_phn = sheet.cell(row= i, column=3)
                        if cell_name.value.lower() == user_name.lower():
                                cell_phn.value = true_phonetices.lower()
                                wb_obj.save(students)
                        tts.say(user_name)
                        tts.runAndWait()
                        break
            break